import json
import authentication_tokens
import environment_url
import openpyxl
from pathlib import Path

orders = []

def dataFile():
    print("\n")
    xlsx = Path('data','D:\API Automation\missing_info_orders.xlsx')
    d1 = openpyxl.load_workbook(xlsx)
    sheet = d1.active
    row_name = []
    phone_num = []
    for column in sheet.iter_rows(1,sheet.max_row):
        row_name.append(column[0].value)
        # phone_num.append(column[1].value)
    cont = 0
    # global phone
    # for j in phone_num:
    #     phone = j
    #     UpdatePhoneNumber(j)

    global orderId
    for i in row_name:
        if i not in orders:
            orderId = i
            orders.append(i)
            cont += 1
            get_Order(i)
        

def get_Order(ord):
    global order
    global orderid
    orderid = str(ord)
    url = environment_url.GetOrder_URL+orderid
    authentication_tokens.parent()
    header = {
        "content-type":"application/json",
        "authorization":"Bearer "+authentication_tokens.parentToken,
        "Organization":"BB-NetJets"
    }
    response = authentication_tokens.requests.request("Get",url,headers = header)
    data = response.text
    order = json.loads(data)
    UpdatePhoneNumber()
        

def UpdatePhoneNumber():
    # Update Phone Number
    order["data"]["CustomerPhone"] = "9142876760"

    global saveData
    saveData = order["data"]
    save_Order()


def save_Order():
    url = environment_url.SaveOrder_URL
    authentication_tokens.parent()
    header = {
        "content-type":"application/json",
        "authorization":"Bearer "+authentication_tokens.parentToken,
        "Organization":"BB-NetJets"
    }
    payload = saveData
    payload = json.dumps(payload)
    authentication_tokens.requests.request("POST",url,headers = header,data = payload)
    print(f"{orderId} - Order updated")

dataFile()